import os, openpyxl
from openpyxl.styles import Alignment
#os.chdir('D:\PythonExcel')
wb = openpyxl.load_workbook('sudoku_board.xlsx')
sheet = wb.active
emptyCell = None
#Scanning for Empty Cells
working = True
while working == True:
	sumTotal = 0
	for columnScan in range(1, 10): #Colunmn Scan
		for rowScan in range(1, 10): #Row Scan
			#possibleValues should always be reset before beginning each search operation
			possibleValues = [1, 2, 3, 4, 5, 6, 7, 8, 9]
			cell = sheet.cell(row = rowScan, column = columnScan)
			#Breaking out of the While Loop
			if cell.value != None:
				sumTotal = sumTotal + int(cell.value)
			if sumTotal == 405: #The total will reach 405, once the puzzle is complete
				working = False

			if cell.value == None: #once an empty cell is found, the loop enters the conditional statement
				emptyCell = cell.coordinate #coordinates will help determine which row and colmn has to be scanned

				#Extracting Row and Colmn index from the coordinate
				stringCoordinate = openpyxl.utils.coordinate_from_string(emptyCell) # e.g. returns ['A',4]
				columnIndex = openpyxl.utils.column_index_from_string(stringCoordinate[0]) # e.g. returns 1 for col 'A'
				rowIndex = stringCoordinate[1]

				#Loop starts only if the cell is empty
				for c in range(1, 10): #fixed range, because it always strats at 1 and ends at 10
					cell = sheet.cell(row = rowIndex, column = c)
					coordinate = cell.coordinate
					#print("Column Scan " + str(cell.value))
					try:
						possibleValues.remove(cell.value)
					except: #Catching ValueError when the code tries to remove nonexistent value
						ValueError
				for r in range(1, 10):
					cell = sheet.cell(row = r, column = columnIndex)
					coordinate = cell.coordinate
					#print("Row Scan " + str(cell.value))
					try:
						possibleValues.remove(cell.value)
					except: #Catching ValueError when the code tries to remove nonexistent value
						ValueError
				#Placing the cell on the top-left corner of any 3 x 3 grid		
				if rowIndex < 4:
					gridRowIndex = 1
				elif rowIndex < 7:
					gridRowIndex = 4
				else:
					gridRowIndex = 7

				if columnIndex < 4:
					gridColumnIndex = 1
				elif columnIndex < 7:
					gridColumnIndex = 4
				else:
					gridColumnIndex = 7

				for gridH in range(gridColumnIndex, gridColumnIndex + 3):
					for gridV in range(gridRowIndex, gridRowIndex + 3):
						cell = sheet.cell(row = gridV, column = gridH)
						try:
							possibleValues.remove(cell.value)
						except: #Catching ValueError when the code tries to remove nonexistent value
							ValueError
				#If only 1 element remaining, assign value to the empty cell
				print("Possible Values for " + emptyCell + "=" + str(len(possibleValues)))
				if len(possibleValues) == 1:
					sheet[emptyCell] = possibleValues[0]
					sheet[emptyCell].alignment = Alignment(horizontal = 'center', vertical = 'center' )
#Finally, save the data in the working memory to disk
wb.save('sudoku_board.xlsx')